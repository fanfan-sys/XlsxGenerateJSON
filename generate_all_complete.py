from typing import Any
import openpyxl
import re
import os
import shutil
import glob
import json
from itertools import product


def main():
    print("=== 开始批量生成所有场景 ===")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)

    output_dir = '已完成语料'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建文件夹: {output_dir}")

    template_dir = '模板'
    all_template_files = glob.glob(os.path.join(template_dir, '【训练语料】*.xlsx'))
    all_template_files = [f for f in all_template_files if '完成版' not in f]

    print(f"\n找到 {len(all_template_files)} 个模板文件")

    success_count = 0
    fail_count = 0

    for template_file in all_template_files:
        scene_name = os.path.basename(template_file).replace('【训练语料】', '').replace('.xlsx', '')

        try:
            process_single_scene(scene_name, output_dir, project_dir)
            success_count += 1
        except Exception as e:
            print(f"处理失败: {scene_name} - {str(e)}")
            fail_count += 1

    print(f"\n=== 生成完成 ===")
    print(f"成功: {success_count}")
    print(f"失败: {fail_count}")

    print(f"\n=== 移动现有完成版文件到文件夹 ===")
    move_existing_complete_files(output_dir)

    print(f"\n全部完成！请查看 '{output_dir}' 文件夹")


def process_single_scene(scene_name, output_dir, project_dir):
    template_dir = '模板'
    template_file = os.path.join(template_dir, f'【训练语料】{scene_name}.xlsx')

    if not os.path.exists(template_file):
        print(f"跳过: 模板文件不存在 - {template_file}")
        return

    safe_scene_name = re.sub(r'[\\/*?:"<>|]', '_', scene_name)
    output_file = os.path.join(output_dir, f'【训练语料】{safe_scene_name}【完成版】.json')

    if os.path.exists(output_file):
        print(f"跳过: 已存在 - {scene_name}")
        return

    print(f"处理: {scene_name}")

    wb_template = openpyxl.load_workbook(template_file)

    single_slot_names = []
    single_slot_vals = {}
    combined_slot_groups = []

    for sheet_name in wb_template.sheetnames:
        if sheet_name not in ['原始信息', '场景ID']:
            if '+' in sheet_name:
                sub_slot_names = sheet_name.split('+')
                ws = wb_template[sheet_name]

                combined_values = []
                for row in ws.iter_rows(values_only=True):
                    if row:
                        row_vals = []
                        valid_row = True
                        for i, sub_slot in enumerate(sub_slot_names):
                            if i < len(row) and row[i]:
                                row_vals.append(str(row[i]))
                            else:
                                valid_row = False
                                break
                        if valid_row:
                            combined_values.append(row_vals)

                if combined_values:
                    combined_slot_groups.append({
                        'slot_names': sub_slot_names,
                        'values': combined_values
                    })
            else:
                single_slot_names.append(sheet_name)
                single_slot_vals[sheet_name] = []
                ws = wb_template[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    if row and row[0]:
                        single_slot_vals[sheet_name].append(str(row[0]))

    all_slot_names = []
    for name in single_slot_names:
        all_slot_names.append(name)
    for group in combined_slot_groups:
        for name in group['slot_names']:
            all_slot_names.append(name)

    ws_template_orig = wb_template['原始信息']
    rows_orig = list(ws_template_orig.iter_rows(values_only=True))
    first_row = rows_orig[1]
    scene_id = str(first_row[11]).strip() if first_row[11] else ''
    action_val = str(first_row[2]) if first_row[2] else ''
    target_val = str(first_row[3]) if first_row[3] else ''

    input_format = str(first_row[12]).strip() if len(first_row) > 12 and first_row[12] else ''
    keyword_format = str(first_row[13]).strip() if len(first_row) > 13 and first_row[13] else ''

    single_slot_values_list = [single_slot_vals.get(name, []) for name in single_slot_names]

    if single_slot_values_list:
        single_slot_combinations = list(product(*single_slot_values_list))
    else:
        single_slot_combinations = [()]

    if combined_slot_groups:
        combined_combinations = list(product(*[group['values'] for group in combined_slot_groups]))
    else:
        combined_combinations = [()]

    all_rows = []
    for single_comb in single_slot_combinations:
        for combined_comb in combined_combinations:
            slot_dict = {}
            for i, name in enumerate(single_slot_names):
                slot_dict[name] = single_comb[i]

            for g_idx, group in enumerate(combined_slot_groups):
                group_values = combined_comb[g_idx]
                for s_idx, slot_name in enumerate(group['slot_names']):
                    slot_dict[slot_name] = group_values[s_idx]

            input_txt = build_input_text(scene_name, slot_dict, all_slot_names, input_format)
            keyword_txt = build_keyword_text(slot_dict, all_slot_names, keyword_format, action_val, input_txt)

            row_obj = build_row_object(input_txt, scene_id, action_val, target_val, all_slot_names, slot_dict,
                                       keyword_txt)
            all_rows.append(row_obj)

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)

    print(f"  保存: {output_file}")


def build_input_text(scene_name, slot_dict, slot_names, input_format=''):
    if input_format:
        text = input_format
        for name in slot_names:
            if name in slot_dict:
                val = slot_dict[name]
                placeholder = '{' + name + '}'
                text = text.replace(placeholder, val)
        text = text.replace('{', '').replace('}', '')
        return text

    text = scene_name
    for name in slot_names:
        if name in slot_dict:
            val = slot_dict[name]
            if name == 'params_opt':
                continue
            if val == 'default':
                continue
            text = text.replace('XX', val, 1)
            text = text.replace('xx', val, 1)
    return text


def build_input_text_short(scene_name, slot_dict, slot_names, input_format=''):
    if input_format:
        text = input_format
        for name in slot_names:
            if name in slot_dict:
                val = slot_dict[name]
                placeholder = '{' + name + '}'
                text = text.replace(placeholder, val)
        text = text.replace('{', '').replace('}', '')
        return text

    text = scene_name

    for name in slot_names:
        if name in slot_dict:
            val = slot_dict[name]
            if name == 'params_opt':
                continue
            if val == 'default':
                continue
            text = text.replace('XX', val, 1)
            text = text.replace('xx', val, 1)

    if '电压状态' in text:
        text = text.replace('电压状态', '')
    elif '主变状态' in text:
        text = text.replace('主变状态', '')
    elif '变压器' in text and '状态' in text:
        text = text.replace('变压器', '').replace('状态', '')
    elif '厂站图' in text:
        text = text.replace('厂站图', '')
    elif '应用画面' in text:
        text = text.replace('应用画面', '')
    elif '刀闸' in text:
        text = text.replace('刀闸', '')
    elif '开关' in text:
        text = text.replace('开关', '')
    elif '变压器' in text:
        text = text.replace('变压器', '')
    elif '机组' in text:
        text = text.replace('机组', '')
    elif '母线' in text:
        text = text.replace('母线', '')
    elif '线路' in text:
        text = text.replace('线路', '')

    return text


def build_keyword_text(slot_dict, slot_names, keyword_format, action_val, input_txt):
    if keyword_format:
        text = keyword_format
        for name in slot_names:
            if name in slot_dict:
                val = slot_dict[name]
                placeholder = '{' + name + '}'
                text = text.replace(placeholder, val)
        text = text.replace('{', '').replace('}', '')
        return text

    keyword_txt = input_txt.replace(action_val, '')
    return keyword_txt


def build_row_object(input_txt, scene_id, action_val, target_val, slot_names, slot_dict, keyword_txt):
    params = {}
    for slot_name in slot_names:
        params[slot_name] = slot_dict[slot_name]

    output_obj = {
        "matchInfo": {
            "id": scene_id
        },
        "semantic": {
            "slots": {
                "action": {
                    "name": action_val
                },
                "target": {
                    "name": target_val,
                    "params": params
                }
            }
        }
    }

    output_str = json.dumps(output_obj, ensure_ascii=False)

    row_obj = {
        "instruction": "",
        "input": input_txt,
        "output": output_str,
        "keyword": keyword_txt
    }
    return row_obj


def build_row(input_txt, scene_id, action_val, target_val, slot_names, slot_dict, keyword_txt):
    parts = []
    parts.append('{"instruction": "","input": "')
    parts.append(input_txt)
    parts.append('","output": "{\\"matchInfo\\":{\\"id\\":\\"')
    parts.append(scene_id)
    parts.append('\\"},\\"semantic\\":{\\"slots\\":{\\"action\\":{\\"name\\":\\"')
    parts.append(action_val)
    parts.append('\\"},\\"target\\":{\\"name\\":\\"')
    parts.append(target_val)
    parts.append('\\",\\"params\\":{')

    first = True
    for slot_name in slot_names:
        if not first:
            parts.append(',')
        first = False
        parts.append('\\"')
        parts.append(slot_name)
        parts.append('\\":\\"')
        parts.append(slot_dict[slot_name])
        parts.append('\\"')

    parts.append('}}}}}","keyword": "')
    parts.append(keyword_txt)
    parts.append('"}')
    parts.append(',')

    return parts


def move_existing_complete_files(output_dir):
    complete_files = glob.glob('【训练语料】*【完成版】*.json')

    moved_count = 0
    for f in complete_files:
        try:
            dest = os.path.join(output_dir, os.path.basename(f))
            if not os.path.exists(dest):
                shutil.move(f, dest)
                moved_count += 1
        except Exception as e:
            print(f"移动失败: {f} - {str(e)}")

    print(f"移动了 {moved_count} 个现有完成版文件")


if __name__ == '__main__':
    main()
